import pandas as pd
import os

# ExcelScheduleManager
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.filters import FilterColumn, Filters
import datetime as dt
from typing import List, Dict
from collections import defaultdict
from csp import University

def createCSV(gen_dir: str = './GoodwingTimetabler/UniversityInstance/'):
    
    xlsx_path = f'{gen_dir}University.xlsx'

    # Ensure the CSV directory exists
    os.makedirs(f'{gen_dir}csv', exist_ok=True)
    
    # Define standard sheets to process
    standard_sheets = ['University', 'Timeslots', 'Promotions', 'Subjects', 'Teachers', 'Rooms']
    
    # Process standard sheets
    wb = load_workbook(xlsx_path)
    for sheet_name in standard_sheets:
        if sheet_name in wb.sheetnames:
            print(f"Processing sheet: {sheet_name}")
            df = pd.read_excel(xlsx_path, sheet_name)
            df.to_csv(f'{gen_dir}csv/{sheet_name}.csv', index=False)
    
    # Process teacher availability if it exists
    if 'TeacherAvailability' in wb.sheetnames:
        print("Processing teacher availability data")
        df = pd.read_excel(xlsx_path, 'TeacherAvailability')
        df.to_csv(f'{gen_dir}csv/TeacherAvailability.csv', index=False)
    else:
        print("No teacher availability data")


def init_template(gen_dir: str = './GoodwingTimetabler/UniversityInstance/', force_reset=True):
    """
    Initialize or reset the Excel template file with all required sheets,
    including the teacher availability sheet.
    
    Parameters:
    - gen_dir: str | Directory to create the template in
    - force_reset: bool | If True, always create a new template even if one exists
    """
    output_path = f'{gen_dir}University.xlsx'
    
    # Create directory if it doesn't exist
    os.makedirs(gen_dir, exist_ok=True)
    
    # Check if file exists and we need to overwrite
    if not os.path.exists(output_path) or force_reset:
        print(f"Creating new template at {output_path}")
        wb = Workbook()
        
        # Remove the default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create University sheet with basic settings
        univ_sheet = wb.create_sheet(title='University')
        univ_sheet.append(['Setting', 'Value'])
        univ_sheet.append(['Name', 'ESILV'])
        univ_sheet.append(['Start day', 6])
        univ_sheet.append(['Start month', 1])
        univ_sheet.append(['Start year', 2025])
        univ_sheet.append(['Days', 14])
        
        # Create Timeslots sheet
        ts_sheet = wb.create_sheet(title='Timeslots')
        ts_sheet.append(['StartH', 'StartMin', 'EndH', 'EndMin'])
        ts_sheet.append([8, 15, 9, 45])
        ts_sheet.append([10, 0, 11, 30])
        ts_sheet.append([11, 45, 13, 15])
        ts_sheet.append([13, 30, 15, 0])
        ts_sheet.append([15, 15, 16, 45])
        ts_sheet.append([17, 0, 18, 30])
        ts_sheet.append([18, 45, 20, 15])
        
        # Create Rooms sheet
        rooms_sheet = wb.create_sheet(title='Rooms')
        rooms_sheet.append(['Name', 'Type'])
        rooms_sheet.append(['L101', 'default'])
        rooms_sheet.append(['L102', 'default'])
        rooms_sheet.append(['L103', 'default'])
        rooms_sheet.append(['Online', 'default'])
        
        # Create Promotions sheet
        promo_sheet = wb.create_sheet(title='Promotions')
        promo_sheet.append(['A1', 'A2'])
        promo_sheet.append(['TDA', 'TDA'])
        
        # Create Subjects sheet
        subj_sheet = wb.create_sheet(title='Subjects')
        subj_sheet.append(['Id', 'Name', 'Promotion', 'Hours', 'Color'])
        # Add some sample subjects
        subj_sheet.append(['UNI011', 'Basic Maths', 'A1', 12, 'FF5733'])
        subj_sheet.append(['UNI012', 'Basic Physics', 'A1', 15, 'FF33A8'])
        subj_sheet.append(['UNI013', 'Basic Informatics', 'A1', 12, '33FF57'])
        subj_sheet.append(['UNI021', 'Advanced Maths', 'A2', 12, 'DC143C'])
        subj_sheet.append(['UNI022', 'Advanced Physics', 'A2', 15, '8A2BE2'])
        subj_sheet.append(['UNI023', 'Advanced Informatics', 'A2', 12, '00FA9A'])
        
        # Create Teachers sheet
        teach_sheet = wb.create_sheet(title='Teachers')
        teach_sheet.append(['Idt', 'First Name', 'Last Name', 'Subjects (séparés d\'un \'-\')'])
        # Add some sample teachers
        teach_sheet.append([0, 'Henri', 'Barbeau', 'UNI011-UNI021'])
        teach_sheet.append([1, 'Timothé', 'Solé', 'UNI011-UNI021'])
        teach_sheet.append([2, 'Renaud', 'Cerfbeer', 'UNI012-UNI022'])
        teach_sheet.append([3, 'Christiane', 'Brunelle', 'UNI013-UNI023'])
        teach_sheet.append([4, 'Constantin', 'Poussin', 'UNI012-UNI022'])
        teach_sheet.append([5, 'Maurice', 'Vannier', 'UNI013-UNI023'])
        
        # Save the workbook
        wb.save(output_path)
        print(f"Created basic template at {output_path}")
    else:
        print(f"Using existing template at {output_path}")
    
    # Now add/update the availability sheet with the same sample data
    create_availability_template(output_path, force_reset)

def create_availability_template(output_path: str = './GoodwingTimetabler/UniversityInstance/University.xlsx', force_reset=True):
    """
    Creates or updates the University.xlsx template to include a TeacherAvailability sheet.
    
    Parameters:
    - output_path: str | Path to the Excel file
    - force_reset: bool | If True, recreate the sheet even if it exists
    """
    try:
        # Try to load the existing workbook
        wb = load_workbook(output_path)
        
        # Check if TeacherAvailability sheet already exists
        if 'TeacherAvailability' in wb.sheetnames:
            if force_reset:
                # Remove the existing sheet
                del wb['TeacherAvailability']
                print("Recreating TeacherAvailability sheet")
            else:
                print("TeacherAvailability sheet already exists in the template.")
                return
        
        # Load teacher and timeslot data to create the availability sheet
        teachers_sheet = wb['Teachers'] if 'Teachers' in wb.sheetnames else None
        timeslots_sheet = wb['Timeslots'] if 'Timeslots' in wb.sheetnames else None
        
        if not teachers_sheet or not timeslots_sheet:
            print("Could not find Teachers or Timeslots sheets. Cannot create availability template.")
            return
        
        # Create the new availability sheet
        avail_sheet = wb.create_sheet(title='TeacherAvailability')
        
        # Get time ranges from Timeslots sheet
        time_ranges = []
        for row in timeslots_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if row[0] is not None and row[1] is not None and row[2] is not None and row[3] is not None:
                start_time = f"{int(row[0])}:{str(int(row[1])).zfill(2)}"
                end_time = f"{int(row[2])}:{str(int(row[3])).zfill(2)}"
                time_ranges.append(f"{start_time}-{end_time}")
        
        # Create column headers for each day and time slot
        days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
        
        # First column is teacher ID
        avail_sheet.cell(row=1, column=1, value="TeacherId")
        
        # Create headers for each day and time slot
        col = 2
        for day in days:
            for time_range in time_ranges:
                header = f"{day}_{time_range}"
                avail_sheet.cell(row=1, column=col, value=header)
                col += 1
        
        # Add teachers from Teachers sheet
        row = 2
        for teacher_row in teachers_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
            if teacher_row[0] is not None:  # Teacher ID
                avail_sheet.cell(row=row, column=1, value=str(teacher_row[0]))
                
                # Set all slots to 1 (available) by default
                for col in range(2, 2 + len(days) * len(time_ranges)):
                    # Create a pattern of availability - for demonstration
                    # Monday-Friday: Available for all slots
                    # Saturday: Only morning slots
                    # Sunday: No availability
                    
                    # Calculate which day and slot this column represents
                    header_idx = col - 2  # 0-based index in the flattened day/time array
                    day_idx = header_idx // len(time_ranges)  # Which day (0-6)
                    time_idx = header_idx % len(time_ranges)  # Which timeslot (0-6)
                    
                    # Set availability based on the pattern
                    if day_idx < 5:  # Mon-Fri
                        value = 1
                    elif day_idx == 5:  # Sat
                        value = 1 if time_idx < 3 else 0  # Only morning slots
                    else:  # Sun
                        value = 0
                    
                    avail_sheet.cell(row=row, column=col, value=value)
                
                row += 1
        
        # Add formatting
        for col in range(1, col):
            column_letter = get_column_letter(col)
            avail_sheet.column_dimensions[column_letter].width = 15 if col == 1 else 12
        
        # Add instructions at the top
        instructions = avail_sheet.cell(row=row+1, column=1, value="Instructions: 1=Available, 0=Unavailable")
        instructions.font = Font(bold=True)
        
        # Save the updated workbook
        wb.save(output_path)
        print(f"TeacherAvailability sheet added to {output_path}")
        
    except Exception as e:
        print(f"Error creating TeacherAvailability template: {e}")

class ExcelScheduleManager:
    def __init__(self, university: University, generated_courses):
        self.university = university
        self.courses = generated_courses
        self.wb = Workbook()
        self.time_slots = university.time_ranges
        
    def format_time(self, time_obj):
        return time_obj.strftime("%H:%M")

    def get_week_number(self, date):
        return (date - self.university.timeslots[0].day).days // 7 + 1

    def setup_sheet_structure(self, ws):
        # Headers
        headers = ['Week', 'Day', 'Time Slot', 'Subject', 'Teacher', 'Room']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add filters
        ws.auto_filter.ref = f'A1:F1'
        
        # Set column widths
        dim_holder = DimensionHolder(worksheet=ws)
        
        widths = [10, 15, 20, 30, 25, 15]  # Adjusted widths for each column
        for col, width in enumerate(widths, 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=width)
        
        ws.column_dimensions = dim_holder

    def add_course_to_sheet(self, ws, row, course, week_num):
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_of_week = day_names[course.timeslot.day.weekday()]
        
        # Data to insert
        data = [
            week_num,
            day_of_week,
            f"{self.format_time(course.timeslot.start)} - {self.format_time(course.timeslot.end)}",
            course.subject.name,
            f"{course.teacher.first_name} {course.teacher.last_name}",
            course.room.name
        ]
        
        # Insert data
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add color to subject cell
            if col == 4:  # Subject column
                cell.fill = PatternFill(start_color=course.subject.color, 
                                      end_color=course.subject.color, 
                                      fill_type="solid")
                # Adjust font color for better readability
                brightness = sum(int(course.subject.color[i:i+2], 16) for i in (0, 2, 4)) / 3
                cell.font = Font(color="FFFFFF" if brightness < 128 else "000000")

        # Add border
        border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'),
                       top=Side(style='thin'),
                       bottom=Side(style='thin'))
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

    def create_group_schedule(self, group_name, group_courses):
        ws = self.wb.create_sheet(title=f"Group_{group_name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(group_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_teacher_schedule(self, teacher, teacher_courses):
        ws = self.wb.create_sheet(title=f"Teacher_{teacher.last_name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(teacher_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_room_schedule(self, room, room_courses):
        ws = self.wb.create_sheet(title=f"Room_{room.name}")
        self.setup_sheet_structure(ws)
        
        row = 2
        for course in sorted(room_courses, 
                           key=lambda x: (self.get_week_number(x.timeslot.day), 
                                        x.timeslot.day, 
                                        x.timeslot.start)):
            week_num = self.get_week_number(course.timeslot.day)
            self.add_course_to_sheet(ws, row, course, week_num)
            row += 1

    def create_statistics_sheet(self):
        ws = self.wb.create_sheet(title="Statistics")
        
        # Style for headers
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        
        # Group Statistics
        ws['A1'] = "Group Statistics"
        ws['A1'].font = header_style
        
        headers = ['Group', 'Total Hours', 'Subjects', 'Teachers']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        row = 3
        group_stats = defaultdict(lambda: {'hours': 0, 'subjects': set(), 'teachers': set()})
        
        for course in self.courses:
            stats = group_stats[course.group.name]
            duration = (course.timeslot.end.hour - course.timeslot.start.hour + 
                       (course.timeslot.end.minute - course.timeslot.start.minute) / 60)
            stats['hours'] += duration
            stats['subjects'].add(course.subject.name)
            stats['teachers'].add(f"{course.teacher.first_name} {course.teacher.last_name}")
        
        for group, stats in group_stats.items():
            ws.cell(row=row, column=1, value=group)
            ws.cell(row=row, column=2, value=f"{stats['hours']:.1f}")
            ws.cell(row=row, column=3, value=len(stats['subjects']))
            ws.cell(row=row, column=4, value=len(stats['teachers']))
            row += 1
        
        # Teacher Statistics
        row += 2
        ws.cell(row=row, column=1, value="Teacher Statistics").font = header_style
        row += 1
        
        headers = ['Teacher', 'Total Hours', 'Groups', 'Subjects']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_style
            cell.fill = header_fill
        
        row += 1
        teacher_stats = defaultdict(lambda: {'hours': 0, 'groups': set(), 'subjects': set()})
        
        for course in self.courses:
            teacher = f"{course.teacher.first_name} {course.teacher.last_name}"
            stats = teacher_stats[teacher]
            duration = (course.timeslot.end.hour - course.timeslot.start.hour + 
                       (course.timeslot.end.minute - course.timeslot.start.minute) / 60)
            stats['hours'] += duration
            stats['groups'].add(course.group.name)
            stats['subjects'].add(course.subject.name)
        
        for teacher, stats in teacher_stats.items():
            ws.cell(row=row, column=1, value=teacher)
            ws.cell(row=row, column=2, value=f"{stats['hours']:.1f}")
            ws.cell(row=row, column=3, value=len(stats['groups']))
            ws.cell(row=row, column=4, value=len(stats['subjects']))
            row += 1
        
        # Adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20

    def generate_excel_schedule(self, output_path):
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Group schedules
        group_courses = defaultdict(list)
        for course in self.courses:
            group_courses[course.group.name].append(course)
        
        for group_name, courses in group_courses.items():
            self.create_group_schedule(group_name, courses)
        
        # Teacher schedules
        teacher_courses = defaultdict(list)
        for course in self.courses:
            teacher_courses[course.teacher].append(course)
        
        for teacher, courses in teacher_courses.items():
            self.create_teacher_schedule(teacher, courses)
        
        # Room schedules
        room_courses = defaultdict(list)
        for course in self.courses:
            room_courses[course.room].append(course)
        
        for room, courses in room_courses.items():
            self.create_room_schedule(room, courses)
        
        # Statistics sheet
        self.create_statistics_sheet()
        
        # Save the workbook
        self.wb.save(output_path)
        print(f"\nBasic timetable saved to {output_path}")

    def create_visual_timetable(self, output_path="./Outputs/excel/visual_timetable.xlsx"):
        """
        Creates a visual Excel timetable with days as columns and timeslots as rows.
        Creates separate sheets for each week to handle unique weekly schedules.
        Each entity (group, teacher, room) gets its own set of weekly sheets.
        Includes an organized index page with separate columns for groups, teachers, and rooms.
        
        Parameters:
        - output_path: str | Path to save the Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.hyperlink import Hyperlink
        from openpyxl.drawing.image import Image
        import datetime as dt
        from collections import defaultdict

        wb = Workbook()
        wb.remove(wb.active)  # Remove the default sheet
        
        # Format time ranges for row headers
        time_range_labels = [f"{self.format_time(start)} - {self.format_time(end)}" 
                            for start, end in self.university.time_ranges]
        
        # Day names for column headers
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        # Create border styles
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Calculate the number of weeks in the schedule
        start_date = self.university.timeslots[0].day
        end_date = self.university.timeslots[-1].day
        days_diff = (end_date - start_date).days
        total_weeks = (days_diff // 7) + 1
        
        # Function to get week number for a date (relative to start_date)
        def get_week_number(date):
            days_from_start = (date - start_date).days
            return days_from_start // 7
        
        # Create the index sheet first
        index = wb.create_sheet(title="Index")
        
        # Keep track of all sheets for organized index
        sheet_info = {
            'Groups': [],
            'Teachers': [],
            'Rooms': []
        }
        
        # Function to add navigation button back to index
        def add_back_to_index_button(worksheet):
            # Add a text hyperlink in the top-right corner
            cell = worksheet.cell(row=1, column=8)  # Use H1 cell
            cell.value = "Back to Index"
            cell.hyperlink = "#Index!A1"
            cell.font = Font(color="0563C1", underline="single", bold=True)
            cell.alignment = Alignment(horizontal='right')
        
        # Function to create a visual timetable sheet for a specific week
        def setup_visual_sheet(entity_name, entity_type, week_num, courses_by_timeslot):
            sheet_type = f"{entity_type}s"  # Convert to plural for categorization
            
            # Format sheet name
            sheet_name = f"{entity_type}_{entity_name}_W{week_num+1}"
            # Truncate sheet name if too long (Excel limit is 31 chars)
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:28] + "..."
            
            # Store sheet info for index
            sheet_info[sheet_type].append({
                'name': entity_name,
                'week': week_num + 1,
                'sheet_name': sheet_name
            })
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Set up column widths
            for col in range(1, 9):  # 1 for time column + 7 days + 1 for back button
                width = 20 if col == 1 else 25
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # Set up row heights
            for row in range(1, len(time_range_labels) + 3):
                ws.row_dimensions[row].height = 60
            
            # Add header with week information
            week_start = start_date + dt.timedelta(days=week_num*7)
            week_end = min(week_start + dt.timedelta(days=6), end_date)
            week_header = f"Week {week_num+1}: {week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"
            
            # Add the entity name and week info in cell A1
            ws.merge_cells('A1:G1')
            header_cell = ws.cell(row=1, column=1)
            header_cell.value = f"{entity_type}: {entity_name} - {week_header}"
            header_cell.font = Font(bold=True, size=14)
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
            header_cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            
            # Add back to index button
            add_back_to_index_button(ws)
            
            # Add day headers (row 2)
            ws.cell(row=2, column=1).value = "Time Slot"
            ws.cell(row=2, column=1).font = Font(bold=True)
            ws.cell(row=2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=2, column=1).border = thin_border
            ws.cell(row=2, column=1).fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            for i, day in enumerate(day_names):
                cell = ws.cell(row=2, column=i+2)
                cell.value = day
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Add time slots as row headers
            for i, time_label in enumerate(time_range_labels):
                cell = ws.cell(row=i+3, column=1)
                cell.value = time_label
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            
            # Create empty grid with borders
            for row in range(3, len(time_range_labels) + 3):
                for col in range(2, 9):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Fill in the courses for this week
            for (day_idx, time_idx), course_list in courses_by_timeslot.items():
                if 0 <= day_idx < 7 and 0 <= time_idx < len(time_range_labels):
                    row = time_idx + 3
                    col = day_idx + 2
                    
                    if course_list:  # If there are courses in this timeslot
                        # Format the text for multiple courses (if more than one)
                        if len(course_list) > 1:
                            text = "\n".join([f"{c.subject.name} ({c.room.name})" for c in course_list])
                            cell = ws.cell(row=row, column=col)
                            cell.value = text
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            cell.font = Font(size=8)  # Smaller font for multiple entries
                        else:
                            # Single course, more detailed formatting
                            course = course_list[0]
                            cell = ws.cell(row=row, column=col)
                            cell.value = f"{course.subject.name}\n{course.teacher.first_name} {course.teacher.last_name}\n({course.room.name})"
                            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                            
                            # Color based on subject
                            color = course.subject.color
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                            
                            # Set font color based on background brightness
                            brightness = sum(int(color[i:i+2], 16) for i in (0, 2, 4)) / 3
                            cell.font = Font(color="FFFFFF" if brightness < 128 else "000000")
        
        # Process group schedules - organize by group, week, day, and timeslot
        group_schedules = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for course in self.courses:
            date = course.timeslot.day
            week_num = get_week_number(date)
            day_idx = date.weekday()  # 0 for Monday, 6 for Sunday
            
            # Find the time range index
            time_idx = None
            for i, (start, end) in enumerate(self.university.time_ranges):
                if course.timeslot.start == start and course.timeslot.end == end:
                    time_idx = i
                    break
            
            if time_idx is not None:
                group_schedules[course.group.name][week_num][(day_idx, time_idx)].append(course)
        
        # Create visual timetable for each group and week
        for group_name, weeks in group_schedules.items():
            for week_num, courses_by_timeslot in weeks.items():
                setup_visual_sheet(group_name, "Group", week_num, courses_by_timeslot)
        
        # Process teacher schedules - organize by teacher, week, day, and timeslot
        teacher_schedules = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for course in self.courses:
            teacher_name = f"{course.teacher.last_name}_{course.teacher.first_name}"
            date = course.timeslot.day
            week_num = get_week_number(date)
            day_idx = date.weekday()
            
            time_idx = None
            for i, (start, end) in enumerate(self.university.time_ranges):
                if course.timeslot.start == start and course.timeslot.end == end:
                    time_idx = i
                    break
            
            if time_idx is not None:
                teacher_schedules[teacher_name][week_num][(day_idx, time_idx)].append(course)
        
        # Create visual timetable for each teacher and week
        for teacher_name, weeks in teacher_schedules.items():
            for week_num, courses_by_timeslot in weeks.items():
                setup_visual_sheet(teacher_name, "Teacher", week_num, courses_by_timeslot)
        
        # Process room schedules - organize by room, week, day, and timeslot
        room_schedules = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for course in self.courses:
            date = course.timeslot.day
            week_num = get_week_number(date)
            day_idx = date.weekday()
            
            time_idx = None
            for i, (start, end) in enumerate(self.university.time_ranges):
                if course.timeslot.start == start and course.timeslot.end == end:
                    time_idx = i
                    break
            
            if time_idx is not None:
                room_schedules[course.room.name][week_num][(day_idx, time_idx)].append(course)
        
        # Create visual timetable for each room and week
        for room_name, weeks in room_schedules.items():
            for week_num, courses_by_timeslot in weeks.items():
                setup_visual_sheet(room_name, "Room", week_num, courses_by_timeslot)
        
        # Now build the index page with three separate columns
        index = wb.worksheets[0]  # Get the index sheet we created earlier
        
        # Set up column widths
        index.column_dimensions['A'].width = 25
        index.column_dimensions['B'].width = 10
        index.column_dimensions['C'].width = 5  # Spacer
        index.column_dimensions['D'].width = 25
        index.column_dimensions['E'].width = 10
        index.column_dimensions['F'].width = 5  # Spacer
        index.column_dimensions['G'].width = 25
        index.column_dimensions['H'].width = 10
        
        # Add title
        index.merge_cells('A1:H1')
        title_cell = index.cell(row=1, column=1)
        title_cell.value = "VISUAL TIMETABLE INDEX"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        
        # Add section headers (row 2)
        section_headers = ["GROUPS", "", "", "TEACHERS", "", "", "ROOMS", ""]
        section_colors = ["9BC2E6", "", "", "A9D08E", "", "", "FFD966", ""]
        
        for i, header in enumerate(section_headers):
            cell = index.cell(row=2, column=i+1)
            cell.value = header
            if header:  # Only style non-empty headers
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color=section_colors[i], end_color=section_colors[i], fill_type="solid")
        
        # Add column headers for each section (row 3)
        col_headers = ["Group Name", "Week", "", "Teacher Name", "Week", "", "Room Name", "Week"]
        
        for i, header in enumerate(col_headers):
            cell = index.cell(row=3, column=i+1)
            cell.value = header
            if header:  # Only style non-empty headers
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                cell.border = thin_border
        
        # Sort entries by name and then by week
        for category in ['Groups', 'Teachers', 'Rooms']:
            sheet_info[category].sort(key=lambda x: (x['name'], x['week']))
        
        # Add entries for groups
        row = 4
        for entry in sheet_info['Groups']:
            index.cell(row=row, column=1).value = entry['name']
            index.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            index.cell(row=row, column=1).border = thin_border
            
            index.cell(row=row, column=2).value = f"Week {entry['week']}"
            index.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
            index.cell(row=row, column=2).border = thin_border
            
            # Add hyperlink to the sheet
            index.cell(row=row, column=1).hyperlink = f"#{entry['sheet_name']}!A1"
            index.cell(row=row, column=1).font = Font(color="0563C1", underline="single")
            row += 1
        
        # Add entries for teachers (start at the same row as groups)
        row = 4
        for entry in sheet_info['Teachers']:
            index.cell(row=row, column=4).value = entry['name']
            index.cell(row=row, column=4).alignment = Alignment(horizontal='left', vertical='center')
            index.cell(row=row, column=4).border = thin_border
            
            index.cell(row=row, column=5).value = f"Week {entry['week']}"
            index.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')
            index.cell(row=row, column=5).border = thin_border
            
            # Add hyperlink to the sheet
            index.cell(row=row, column=4).hyperlink = f"#{entry['sheet_name']}!A1"
            index.cell(row=row, column=4).font = Font(color="0563C1", underline="single")
            row += 1
        
        # Add entries for rooms (start at the same row as groups)
        row = 4
        for entry in sheet_info['Rooms']:
            index.cell(row=row, column=7).value = entry['name']
            index.cell(row=row, column=7).alignment = Alignment(horizontal='left', vertical='center')
            index.cell(row=row, column=7).border = thin_border
            
            index.cell(row=row, column=8).value = f"Week {entry['week']}"
            index.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
            index.cell(row=row, column=8).border = thin_border
            
            # Add hyperlink to the sheet
            index.cell(row=row, column=7).hyperlink = f"#{entry['sheet_name']}!A1"
            index.cell(row=row, column=7).font = Font(color="0563C1", underline="single")
            row += 1
        
        # Add instructions at the bottom
        max_row = max(
            4 + len(sheet_info['Groups']),
            4 + len(sheet_info['Teachers']),
            4 + len(sheet_info['Rooms'])
        ) + 2
        
        index.merge_cells(f'A{max_row}:H{max_row}')
        index.cell(row=max_row, column=1).value = "Click on any name to view the corresponding timetable"
        index.cell(row=max_row, column=1).font = Font(italic=True)
        index.cell(row=max_row, column=1).alignment = Alignment(horizontal='center')
        
        # Add university name and generation date
        max_row += 2
        index.merge_cells(f'A{max_row}:H{max_row}')
        generation_date = dt.datetime.now().strftime("%d/%m/%Y %H:%M")
        index.cell(row=max_row, column=1).value = f"{self.university.name} - Generated on {generation_date}"
        index.cell(row=max_row, column=1).font = Font(italic=True)
        index.cell(row=max_row, column=1).alignment = Alignment(horizontal='center')
        
        # Save the workbook
        wb.save(output_path)
        print(f"\nVisual timetable saved to {output_path}")